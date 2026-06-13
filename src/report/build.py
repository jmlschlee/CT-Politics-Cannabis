"""Generate the four outputs:
  * out/tracker.xlsx       — working tracker (House / Senate / Former sheets)
  * out/findings.md / .pdf — ranked CONFIRMED findings + recusals, then leads, then table
  * out/review_queue.csv   — every PROBABLE/REVIEW match + every family lead

The caveat is baked into every output: absence of a match is "No match found,"
not proof of no involvement.
"""
from __future__ import annotations

import csv
import json
import re
import shutil
from datetime import date
from pathlib import Path

from ..analyze import LEGAL_PREAMBLE
from ..config import ROOT
from ..models import Finding, Legislator

# Program identity (the user-facing name).
PROGRAM_NAME = "CTCannabisPoliticalCheck"
DISPLAY_NAME = "CT Cannabis Political Check"
VERSION = "1.2.0"


def app_version() -> str:
    """The released program version (single source of truth)."""
    return VERSION

# Numbered reports are PRESERVED here (never overwritten); the registry tracks the
# next number and survives `make clean` (which only touches out/).
REPORTS_DIR = ROOT / "reports"
REGISTRY_PATH = REPORTS_DIR / "registry.json"

# Reader-facing relationship-tier labels. The internal logic strings
# (CONFIRMED / PROBABLE / POSSIBLE / SURNAME ONLY) stay stable so the
# verified-resolution cache, the matcher, and the test suite are unaffected;
# only the wording shown to a human in the report changes.
DISPLAY_TIER = {
    "CONFIRMED": "VERIFIED",
    "PROBABLE": "HIGH PROBABILITY",
    "POSSIBLE": "POSSIBLE",
    "POSSIBLE/REVIEW": "POSSIBLE",
    "SURNAME ONLY": "UNVERIFIED NAME MATCH",
    "NOT VERIFIED": "UNVERIFIED NAME MATCH",
}


def display_tier(tier: str) -> str:
    """Map an internal relationship tier to its reader-facing label."""
    return DISPLAY_TIER.get((tier or "").upper(), tier or "")


def next_report_number() -> int:
    """Return the next report number (>=1), robust to a lost/cleared registry:
    the next number is one past BOTH the registry counter and the highest
    CTCannabisPoliticalCheck_N.pdf already on disk, so prior reports are never
    overwritten."""
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    reg_next = 1
    if REGISTRY_PATH.exists():
        try:
            reg_next = int(json.loads(REGISTRY_PATH.read_text()).get("next", 1))
        except Exception:  # noqa: BLE001
            reg_next = 1
    on_disk = 0
    for p in REPORTS_DIR.glob(f"{PROGRAM_NAME}_*.pdf"):
        m = re.search(rf"{PROGRAM_NAME}_(\d+)\.pdf$", p.name)
        if m:
            on_disk = max(on_disk, int(m.group(1)))
    return max(reg_next, on_disk + 1)


def _record_report_number(n: int, info: dict) -> None:
    history = []
    if REGISTRY_PATH.exists():
        try:
            history = json.loads(REGISTRY_PATH.read_text()).get("history", [])
        except Exception:  # noqa: BLE001
            history = []
    history.append({"number": n, **info})
    REGISTRY_PATH.write_text(json.dumps(
        {"next": n + 1, "history": history}, indent=2), encoding="utf-8")

CATEGORY_COLUMN = {
    "business": "CONCORD",
    "dcp": "DCP cannabis license",
    "donation": "SEEC campaign finance",
    "lobbyist": "Indirect ties",
    "sfi": "Indirect ties",
}
TRACKER_COLUMNS = [
    "District", "Name", "Party", "Leadership", "Hometown", "First elected",
    "Occupation", "CONCORD", "DCP cannabis license", "SEEC campaign finance",
    "Indirect ties", "Priority", "Findings / citations",
]
_STATUS_RANK = {
    "HIT — see findings": 3,
    "Appearance concern": 2,
    "Unable to verify": 1,
    "No match found": 0,
}
_PRIO_RANK = {"HIGH": 3, "MEDIUM": 2, "LOW": 1, "": 0}
CAVEAT = ("CAVEAT: 'No match found' means no match was found in the queried "
          "sources — NOT proof of no involvement. Any source that could not be "
          "exhaustively queried (e.g. a rate-limited interactive portal) is flagged.")

# Hard disclaimer printed prominently on the cover and in the app: every item is an
# UNVERIFIED LEAD until a human confirms it against the cited primary source, and
# nothing here is a legal conclusion.
DISCLAIMER = (
    "IMPORTANT — READ FIRST: This is a research SCREENING AID, not an accusation, a "
    "legal finding, or proof of wrongdoing. Every name, relationship, donation, vote, "
    "and tier below is a LEAD THAT MUST BE INDEPENDENTLY VERIFIED against the cited "
    "primary source before it is relied on or repeated. A shared name is not proof of "
    "the same person. Nothing here establishes a legal conflict of interest, ethics "
    "violation, or crime — those are determinations only a court or the proper "
    "authority can make. Contributions and lobbying shown are LAWFUL, PUBLICLY "
    "DISCLOSED activity provided for context, not allegations. Use responsibly.")


def _per_member(legislators: list[Legislator], findings: list[Finding]) -> dict:
    """person_id -> {column -> status, priority, citations[]}."""
    by_person: dict[str, dict] = {}
    for leg in legislators:
        by_person[leg.person_id] = {
            "leg": leg,
            "cells": {c: "No match found" for c in
                      ("CONCORD", "DCP cannabis license", "SEEC campaign finance",
                       "Indirect ties")},
            "priority": "",
            "citations": [],
        }
    for f in findings:
        rec = by_person.get(f.person_id)
        if not rec:
            continue
        col = CATEGORY_COLUMN.get(f.category, "Indirect ties")
        if _STATUS_RANK.get(f.status, 0) > _STATUS_RANK.get(rec["cells"][col], 0):
            rec["cells"][col] = f.status
        if _PRIO_RANK.get(f.priority, 0) > _PRIO_RANK.get(rec["priority"], 0):
            rec["priority"] = f.priority
        if f.publishable or f.status in ("HIT — see findings", "Appearance concern"):
            tag = "" if f.publishable else " (UNVERIFIED — review)"
            rec["citations"].append(f"[{f.category}] {f.status}{tag}: " +
                                    "; ".join(f.citations))
    return by_person


def _leadership(leg: Legislator) -> str:
    return "General Law/Judiciary member" if leg.flags_relevant_committee else ""


def _town_map_sheet(wb, municipal) -> None:
    """One row per (town, cannabis business) — the §7 Town map sheet."""
    from openpyxl.styles import Font
    ws = wb.create_sheet("Town map")
    cols = ["Town", "Cannabis business", "Facility address", "Approving body",
            "Vote", "Officials in office at approval", "Town counsel firm (reps cannabis?)",
            "Family-representation flags", "Local-entity vendor flags",
            "Legislative overlay", "Confirmed", "Unconfirmed", "Unsupported", "Context"]
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True)
    for d in municipal.dossiers:
        fam = "; ".join(f"{x.subject_name} ({x.classification})" for x in d.connections
                        if x.connection_type == "official_family_rep")
        vend = "; ".join(f"{x.subject_name} ({x.classification})" for x in d.connections
                         if x.connection_type == "vendor_contractor")
        overlay = "; ".join(x.subject_name for x in d.connections
                            if x.connection_type == "legislative_overlay")
        firmflag = "; ".join(x.subject_name for x in d.connections
                             if x.subject_kind == "firm")
        ws.append([
            d.town, d.operator, d.facility.address, d.facility.approval_body,
            d.facility.approval_vote, "", firmflag or "(none surfaced)",
            fam or "(none)", vend or "(none)", overlay or "(none)",
            len(d.by_class("CONFIRMED")), len(d.by_class("UNCONFIRMED")),
            len(d.by_class("UNSUPPORTED")), len(d.by_class("CONTEXT")),
        ])
    ws.freeze_panes = "A2"


def write_tracker(path: Path, legislators: list[Legislator],
                  findings: list[Finding], run_date: str, municipal=None) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    pm = _per_member(legislators, findings)
    wb = Workbook()
    wb.remove(wb.active)

    def add_sheet(title: str, members: list[Legislator]):
        ws = wb.create_sheet(title[:31])
        ws.append(TRACKER_COLUMNS)
        for c in ws[1]:
            c.font = Font(bold=True)
        for leg in sorted(members, key=lambda l: (l.chamber or "", str(l.district))):
            rec = pm[leg.person_id]
            ws.append([
                leg.district, leg.full_name, leg.party, _leadership(leg),
                leg.hometown, leg.first_elected or "", leg.occupation,
                rec["cells"]["CONCORD"], rec["cells"]["DCP cannabis license"],
                rec["cells"]["SEEC campaign finance"], rec["cells"]["Indirect ties"],
                rec["priority"] or "—",
                " | ".join(rec["citations"]) if rec["citations"] else "",
            ])
        ws.freeze_panes = "A2"

    house = [l for l in legislators if not l.is_former and l.chamber == "House"]
    senate = [l for l in legislators if not l.is_former and l.chamber == "Senate"]
    former = [l for l in legislators if l.is_former]
    add_sheet("House", house)
    add_sheet("Senate", senate)
    add_sheet("Former members", former)

    info = wb.create_sheet("How to use")
    notes = [
        ["CT Legislature Cannabis Conflict-of-Interest Screening — Tracker"],
        [f"Generated: {run_date}"],
        [""],
        ["This is a SCREENING AID FOR HUMANS, not an automated accusation engine."],
        [CAVEAT],
        [""],
        ["Column meaning:"],
        ["  CONCORD = CT Secretary of the State business registry (LLC principals)"],
        ["  DCP cannabis license = DCP licenses + INDIVIDUAL credentials (backer, key employee)"],
        ["  SEEC campaign finance = cannabis-affiliated contributions to the member's committees"],
        ["  Indirect ties = relative-lobbyist leads + SFI spouse/family employer leads"],
        ["  Priority = HIGH where a CONFIRMED direct stake or a documented recusal exists"],
        [""],
        ["Every PROBABLE/REVIEW match and every family/spouse lead is in out/review_queue.csv"],
        ["with its source citations, for human sign-off. Nothing there is a finding until verified."],
        [""],
        ["Legal standard: " + LEGAL_PREAMBLE],
    ]
    for row in notes:
        info.append(row)

    if municipal is not None and municipal.dossiers:
        _town_map_sheet(wb, municipal)
    wb.save(str(path))


def write_review_queue(path: Path, review_rows: list[dict]) -> None:
    cols = ["person", "district", "category", "confidence", "status", "ref_label",
            "match_explanation", "is_family_lead", "legal_basis", "source_url"]
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for r in review_rows:
            w.writerow({c: r.get(c, "") for c in cols})


def _findings_sections(findings: list[Finding], recusals: list):
    published = [f for f in findings if f.publishable]
    published.sort(key=lambda f: (-_PRIO_RANK.get(f.priority, 0), f.person_name))
    leads = [f for f in findings if not f.publishable
             and f.status in ("HIT — see findings", "Appearance concern", "Unable to verify")]
    leads.sort(key=lambda f: (f.is_family_lead is False, f.person_name))
    return published, leads


def _dossier_section(municipal) -> list[str]:
    """Per-town dossier in the four-class structure of §4.1 (Simsbury pattern).
    For a large bulk run, a compact facility map plus full dossiers only for towns
    with a substantive (non-context) connection."""
    from ..analyze import MUNICIPAL_POLICY
    L = ["## Municipal layer — host-town facility map\n", f"_{MUNICIPAL_POLICY}_\n"]
    L.append(f"_{len(municipal.dossiers)} cannabis facilities across "
             f"{municipal.counts.get('host_towns', 0)} host towns. Town-level "
             f"official/family/vendor cross-references need per-town sources (no "
             f"statewide bulk API — a flagged coverage gap)._\n")
    L.append("| Town | Operator | License type |")
    L.append("|---|---|---|")
    for d in sorted(municipal.dossiers, key=lambda d: (d.town, d.operator)):
        L.append(f"| {d.town} | {d.operator} | {d.facility.license_type or ''} |")
    L.append("")
    substantive = [d for d in municipal.dossiers
                   if any(c.classification != "CONTEXT" for c in d.connections)]
    if not substantive:
        return L
    L.append("## Per-town dossiers (towns with a substantive connection)\n")
    CLASS_LABEL = {
        "CONFIRMED": "CONFIRMED (well-sourced)",
        "UNCONFIRMED": "UNCONFIRMED — must NOT be asserted as fact",
        "UNSUPPORTED": "UNSUPPORTED — checked and not found (negative finding)",
        "CONTEXT": "CONTEXT-ONLY — relevant but not a financial conflict",
    }
    for d in substantive:
        L.append(f"### {d.town} — {d.operator}\n")
        L.append(f"_Facility: {d.facility.address or d.town}; "
                 f"{d.facility.approval_outcome or 'approved'} by "
                 f"{d.facility.approval_body or 'local body'} "
                 f"({d.facility.approval_vote or 'n/a'})._\n")
        for klass in ("CONFIRMED", "UNCONFIRMED", "UNSUPPORTED", "CONTEXT"):
            items = d.by_class(klass)
            L.append(f"**{CLASS_LABEL[klass]}**")
            if not items:
                L.append("- _(none)_")
            for c in items:
                sub = " — **substantial conflict**" if c.substantial_conflict else (
                    " — appearance concern" if c.appearance_concern else "")
                cites = ", ".join(c.citations) if c.citations else "—"
                L.append(f"- [{c.connection_type}] **{c.subject_name}**{sub}: "
                         f"{c.explanation}  \n  _sources: {cites}_")
            L.append("")
    return L


_STATUS_LABEL = {
    "live": "✅ LIVE (queried this run)",
    "cache": "✅ cached (from a prior live run)",
    "fixture": "⚠️ fixture (offline demo data)",
    "unavailable": "⛔ NOT QUERIED — no bulk API",
    "disabled": "⛔ disabled",
    "": "—",
}
# Plain (no-emoji) variant for the PDF — reportlab's base fonts have no emoji glyphs.
_STATUS_PLAIN = {
    "live": "LIVE (queried this run)",
    "cache": "cached (prior live run)",
    "fixture": "fixture (offline demo data)",
    "unavailable": "NOT QUERIED — no bulk API",
    "disabled": "disabled",
    "": "—",
}


def _coverage_md(coverage: dict, municipal_coverage: dict | None) -> list[str]:
    L = ["## Sources queried (coverage)\n",
         "_Absence of a match is 'no match found,' **not** proof of no involvement. "
         "Any source not queried below was not exhaustively checked._\n",
         "| Source | Status | Records |", "|---|---|---|"]
    for label, c in coverage.items():
        L.append(f"| {label} | {_STATUS_LABEL.get(c['status'], c['status'])} | "
                 f"{c['count']} |")
    if municipal_coverage:
        for label, c in municipal_coverage.items():
            L.append(f"| (town) {label} | {_STATUS_LABEL.get(c['status'], c['status'])}"
                     f" | {c['count']} |")
    L.append("")
    gaps = [(lbl, c) for lbl, c in {**coverage, **(municipal_coverage or {})}.items()
            if c["status"] in ("unavailable", "disabled") and c.get("note")]
    if gaps:
        L.append("**Coverage gaps (why a source could not be exhaustively queried):**\n")
        for lbl, c in gaps:
            L.append(f"- **{lbl}** — {c['note']}")
        L.append("")
    return L


def write_findings_md(path: Path, legislators: list[Legislator],
                      findings: list[Finding], recusals: list, run_date: str,
                      counts: dict, municipal=None, coverage: dict | None = None,
                      mode: str = "OFFLINE", leads: list | None = None,
                      network=None) -> None:
    published, unverified_findings = _findings_sections(findings, recusals)
    L = []
    L.append("# CT Legislature — Cannabis Conflict-of-Interest Screening\n")
    L.append(f"_Generated {run_date} · **{mode} run**. Screening aid for humans — "
             f"not an automated accusation engine._\n")
    L.append(f"> {CAVEAT}\n")

    # Completeness verdict (zero matches only OK if every source queried).
    verdict = _completeness_verdict(coverage or {},
                                    getattr(municipal, "coverage", None) if municipal else None)
    if verdict["complete"]:
        L.append("## ✅ Investigation status: COMPLETE — every required source was queried\n")
    else:
        L.append("## ⛔ Investigation status: INCOMPLETE — not all required sources were queried\n")
        L.append("_A 'no match' below is NOT a clearance. Sources still required:_\n")
        L.append("| Required source NOT fully queried | What it would add | Why / status |")
        L.append("|---|---|---|")
        for k, d, n in verdict["missing"]:
            L.append(f"| {k} | {d} | {n} |")
        L.append("")

    # Section 1 — State legislators & cannabis connections (priority leads).
    L.append("## 1. State legislators & cannabis connections (priority leads)\n")
    L.append("_Cannabis-era (2012+) legislators sharing a surname with a resolved "
             "cannabis-business principal/agent — a LEAD (self, relative, or "
             "coincidence) for human verification, never a finding on its own._\n")
    if leads:
        L.append("| Official & office | Party / town | Cannabis principal | Cannabis LLC | "
                 "Role | Confidence | Sim | How connected | Sources |")
        L.append("|---|---|---|---|---|---|---|---|---|")
        for d in leads:
            srcs = " ".join(f"[src]({u})" for u in d.get("source_urls", []))
            L.append(f"| {d['person']} ({d['role']}) | {d['party']}, "
                     f"{d['district_or_town']} | {d['cannabis_person']} | "
                     f"{d['cannabis_entity']} | {d['cannabis_role']} | "
                     f"{d['confidence']}{' (common surname)' if d.get('is_common_surname') else ''} | "
                     f"{d['name_similarity']} | {d['explanation']} | {srcs} |")
        L.append("")
    else:
        L.append("_No legislator-principal surname leads surfaced this run._\n")

    # Section 2 — Ownership network (people behind cannabis LLCs).
    if network is not None and getattr(network, "edges", None):
        ppl = [e for e in network.edges if e.is_person]
        L.append("## 2. Cannabis LLC ownership / principal / registered-agent network\n")
        L.append(f"_Resolved {network.matched_entities} cannabis businesses via "
                 f"{len(network.edges)} ownership edges to {len(ppl)} individuals "
                 f"(home addresses NOT stored); {len(network.unmatched_entities)} "
                 f"businesses unmatched (INCOMPLETE)._\n")
        L.append("| Person | Role | Cannabis business (chain) | City | Filing date | Source |")
        L.append("|---|---|---|---|---|---|")
        seen = set()
        for e in sorted(ppl, key=lambda e: (e.root_entity.lower(), e.person_or_org.lower())):
            k = (e.person_or_org, e.root_entity, e.role)
            if k in seen:
                continue
            seen.add(k)
            chain = e.root_entity + (f" (via {e.parent})" if e.parent and e.parent != e.root_entity else "")
            L.append(f"| {e.person_or_org} | {e.role} | {chain} | {e.business_city} | "
                     f"{e.date} | [src]({e.source_url}) |")
        L.append("")

    if coverage:
        L += _coverage_md(coverage, getattr(municipal, "coverage", None) if municipal else None)
    L.append("## Legal standard\n")
    L.append(LEGAL_PREAMBLE + "\n")

    L.append("## Run summary\n")
    L.append("| metric | count |\n|---|---|")
    for k in ("legislators", "current", "former", "cannabis_entities",
              "cannabis_persons", "contributions", "lobbyists", "sfi",
              "matches", "published", "review_queue", "recusals"):
        L.append(f"| {k} | {counts.get(k, 0)} |")
    L.append("")

    L.append("## Documented recusals on cannabis votes (strongest signal)\n")
    if recusals:
        for r in recusals:
            L.append(f"- **{r.member_name}** ({r.chamber}, {r.date}) — {r.subject}. "
                     f"_{r.snippet}_ — source: [{r.source_name}]({r.source_url})")
    else:
        L.append("_None found in the parsed journals/committee records this run._")
    L.append("")

    L.append("## Confirmed findings (CONFIRMED identity, with citations)\n")
    if published:
        for f in published:
            links = " ".join(f"[[source]]({u})" for u in f.source_urls)
            L.append(f"- **{f.person_name}** — _{f.category}_ — **{f.status}** "
                     f"[{f.priority}] — {f.legal_basis}\n  - {f.explanation}\n  - "
                     f"citations: {', '.join(f.citations)} {links}".rstrip())
    else:
        L.append("_No findings reached the CONFIRMED-and-publishable bar this run._")
    L.append("")

    L.append("## UNVERIFIED LEADS (NOT findings — pending human review)\n")
    L.append("_These are in `review_queue.csv`. They include every probable/possible "
             "match and every family/spouse lead. Do not treat as conflicts until a "
             "human verifies them against an authoritative source._\n")
    if unverified_findings:
        for f in unverified_findings:
            fam = " — FAMILY/SPOUSE LEAD" if f.is_family_lead else ""
            L.append(f"- {f.person_name} — _{f.category}_ — {f.confidence}{fam} — "
                     f"{f.explanation}")
    else:
        L.append("_No leads this run._")
    L.append("")

    pm = _per_member(legislators, findings)

    def _flagged(leg):
        c = pm[leg.person_id]["cells"]
        return any(v != "No match found" for v in c.values())

    flagged = [l for l in legislators if _flagged(l)]
    big = len(legislators) > 200
    shown = flagged if big else legislators
    title = ("## Per-member table — members with at least one match\n"
             if big else "## Full per-member table\n")
    L.append(title)
    if big:
        L.append(f"_The full {len(legislators):,}-member roster is in "
                 f"`tracker.xlsx`. {len(legislators) - len(flagged):,} members had "
                 f"'No match found' across all categories and are omitted here._\n")
    L.append("| District | Name | Party | CONCORD | DCP | SEEC | Indirect | Priority |")
    L.append("|---|---|---|---|---|---|---|---|")
    for leg in sorted(shown, key=lambda l: (l.is_former, l.chamber or "", str(l.district))):
        rec = pm[leg.person_id]
        c = rec["cells"]
        L.append(f"| {leg.district} | {leg.full_name}"
                 f"{' (former)' if leg.is_former else ''} | {leg.party} | "
                 f"{c['CONCORD']} | {c['DCP cannabis license']} | "
                 f"{c['SEEC campaign finance']} | {c['Indirect ties']} | "
                 f"{rec['priority'] or '—'} |")
    if big and not flagged:
        L.append("| — | _(no members matched any cannabis source this run)_ |||||||")
    L.append("")

    if municipal is not None and municipal.dossiers:
        L += _dossier_section(municipal)
    path.write_text("\n".join(L) + "\n", encoding="utf-8")


def _esc(text: str) -> str:
    """Escape text for reportlab's mini-markup."""
    return (str(text or "").replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;"))


def _esc_attr(url: str) -> str:
    """Escape a URL for use inside an href="" attribute."""
    return (str(url or "").replace("&", "&amp;").replace('"', "%22")
            .replace("<", "%3C").replace(">", "%3E"))


class _Refs:
    """Citation registry: turns a URL into a clickable [n] marker and accumulates a
    numbered, clickable References appendix."""
    LINK = "#1a5276"

    def __init__(self):
        self.order: list[str] = []
        self.index: dict[str, int] = {}

    def cite(self, *urls: str) -> str:
        marks = []
        for url in urls:
            if not url:
                continue
            if url not in self.index:
                self.index[url] = len(self.order) + 1
                self.order.append(url)
            n = self.index[url]
            marks.append(f'<a href="{_esc_attr(url)}" color="{self.LINK}">[{n}]</a>')
        return " " + "".join(marks) if marks else ""

    def link(self, url: str, text: str | None = None) -> str:
        return (f'<a href="{_esc_attr(url)}" color="{self.LINK}">'
                f'{_esc(text or url)}</a>')


# Required sources for a COMPLETE investigation. A run is INCOMPLETE unless every
# one was actually queried with usable data (the user's hard requirement).
REQUIRED_SOURCES = [
    ("Legislators (current+historical)", "state roster"),
    ("DCP cannabis licenses", "cannabis businesses"),
    ("Cannabis ownership network (registry principals/agents)", "LLC -> people"),
    ("Campaign finance", "cannabis-linked donations"),
    ("Lobbyists", "cannabis lobbyists/clients"),
    ("Statements of Financial Interests", "spouse/family employer"),
    ("(town) Municipal officials", "local officials"),
    ("(town) Meeting minutes", "zoning/approval votes + recusals"),
    ("(town) Town counsel / law firms", "town attorney cannabis ties"),
]
# Sources still needed that have no public bulk API at all (always a gap until built).
ALWAYS_GAP = [
    ("DCP individual cannabis credentials (backers, key employees)",
     "names tied directly to each license — only on the eLicense roster "
     "(elicense.ct.gov/Lookup/GenerateRoster.aspx, ASP.NET form, not a bulk API; "
     "rank-and-file employees intentionally excluded). The ownership-network "
     "principals/agents below are the integrated substitute."),
    ("CGA cannabis roll-call votes", "per-bill yes/no/abstain + documented recusals "
     "(cga.ct.gov — HTML/PDF, not yet integrated)"),
    ("Local cannabis votes / moratoria / host-community agreements",
     "per-town council & P&Z actions behind each zoning status (town sites/BoardDocs)"),
]


def _completeness_verdict(coverage: dict, municipal_coverage: dict | None) -> dict:
    """COMPLETE only if every required source returned usable live/cached data."""
    merged = dict(coverage or {})
    for k, v in (municipal_coverage or {}).items():
        merged[f"(town) {k}"] = v
    queried, missing = [], []
    for key, desc in REQUIRED_SOURCES:
        c = merged.get(key)
        ok = bool(c) and c.get("status") in ("live", "cache") and c.get("count", 0) >= 0 \
            and c.get("status") not in ("unavailable", "disabled")
        # A live source with 0 records still counts as "queried" (an honest zero).
        if c and c.get("status") in ("live", "cache"):
            queried.append((key, desc, c.get("count", 0)))
        else:
            missing.append((key, desc, (c or {}).get("note", "not integrated")))
    missing += [(k, d, "no public bulk API") for k, d in ALWAYS_GAP]
    return {"complete": len(missing) == 0, "queried": queried, "missing": missing}


def _wrap_table(header_cells, body_rows, col_widths, cell_style, head_style,
                head_bg="#16412b"):
    """A full-width table whose cells WRAP (Paragraph cells), so long LLC names,
    attorney names, URLs and explanations never clip. `body_rows` cells are markup
    strings (already escaped / may contain <a> links)."""
    from reportlab.lib import colors
    from reportlab.platypus import Paragraph, Table, TableStyle

    data = [[Paragraph(_esc(h), head_style) for h in header_cells]]
    for row in body_rows:
        data.append([Paragraph(c if c is not None else "", cell_style) for c in row])
    t = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(head_bg)),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#f3f6f4")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    return t


def _page_footer(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 7)
    canvas.setFillColorRGB(0.4, 0.4, 0.4)
    canvas.drawString(54, 26, f"{DISPLAY_NAME} — screening aid for humans, not an "
                              f"accusation engine")
    canvas.drawRightString(558, 26, f"Page {doc.page}")
    canvas.restoreState()


def write_findings_pdf(path: Path, findings: list[Finding], recusals: list,
                       run_date: str, counts: dict, legislators=None,
                       municipal=None, coverage: dict | None = None,
                       mode: str = "OFFLINE", report_number: int | None = None,
                       leads: list | None = None, network=None,
                       campaign_finance: dict | None = None,
                       lobbying: dict | None = None) -> None:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
        HRFlowable)

    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus import KeepTogether
    styles = getSampleStyleSheet()
    body = ParagraphStyle("body", parent=styles["Normal"], fontSize=9, leading=12)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8, leading=11)
    note = ParagraphStyle("note", parent=small, textColor=colors.HexColor("#7a5b00"))
    # Section headers: CENTERED + UPPERCASE, consistent throughout.
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=13, alignment=TA_CENTER,
                        textColor=colors.HexColor("#16412b"), spaceBefore=12, spaceAfter=4)
    h3 = ParagraphStyle("h3", parent=styles["Heading3"], fontSize=10.5, alignment=TA_CENTER,
                        textColor=colors.HexColor("#1a5276"), spaceBefore=8)

    def _H2(text):  # centered, uppercase section header
        return Paragraph(text.upper(), h2)

    def _H3(text):
        return Paragraph(text.upper(), h3)

    def _section(header_text, *flowables):
        """Keep a section header attached to its first table/flowable (no orphan
        header at a page bottom)."""
        S.append(KeepTogether([header_text] + list(flowables)))
    # Default word-wrap (wraps at spaces, keeps whole words like "CONFIRMED" intact;
    # reportlab still splits a token only if it truly cannot fit the column).
    cell = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7.3, leading=9,
                          splitLongWords=1)
    cellh = ParagraphStyle("cellh", parent=cell, textColor=colors.white,
                           fontName="Helvetica-Bold")
    refs = _Refs()
    PAGE_W = 612 - 108  # letter minus 54pt margins = 504pt usable width

    rep = f" — Report #{report_number}" if report_number else ""
    doc = SimpleDocTemplate(
        str(path), pagesize=letter,
        title=f"{DISPLAY_NAME}{rep}", author=PROGRAM_NAME,
        leftMargin=54, rightMargin=54, topMargin=54, bottomMargin=44)
    S: list = []

    # -- cover / header ---------------------------------------------------
    S.append(Paragraph(f"{DISPLAY_NAME}", styles["Title"]))
    S.append(Paragraph("Connecticut Legislature &amp; Municipal Cannabis "
                       "Conflict-of-Interest Screening", styles["Heading3"]))
    S.append(Paragraph(f"Generated {run_date}"
                       f"{' &nbsp;·&nbsp; <b>Report #' + str(report_number) + '</b>' if report_number else ''}"
                       f" &nbsp;·&nbsp; <b>{_esc(mode)} run</b> &nbsp;·&nbsp; "
                       f"screening aid for humans, <b>not</b> an automated "
                       f"accusation engine", styles["Italic"]))
    S.append(Spacer(1, 6))
    S.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#16412b")))
    S.append(Spacer(1, 6))
    # Prominent verification / no-legal-implications disclaimer box.
    disc = ParagraphStyle("disc", parent=small, textColor=colors.HexColor("#7a1e1e"),
                          borderColor=colors.HexColor("#7a1e1e"), borderWidth=1,
                          borderPadding=6, backColor=colors.HexColor("#fbf3f3"),
                          leading=12)
    S.append(Paragraph("<b>" + _esc(DISCLAIMER) + "</b>", disc))
    S.append(Spacer(1, 6))
    # OFFLINE = synthetic fixtures: say so unmistakably so demo output is NEVER read
    # as real findings about real people. (Live runs use real, sourced data.)
    if mode != "LIVE":
        synth = ParagraphStyle("synth", parent=small, textColor=colors.white,
                               backColor=colors.HexColor("#7a1e1e"), borderPadding=7,
                               leading=12)
        S.append(Paragraph(
            "&#9888; OFFLINE DEMO &#8212; SYNTHETIC FIXTURES. The legislator, "
            "campaign-finance, and lobbyist entries below use FICTIONAL names "
            "(Hallowell, Vance, Aldenberry, Brightwood, &#8230;) to demonstrate the "
            "report format &#8212; they are NOT findings and describe NO real "
            "official's conduct. (The Simsbury / Glassman municipal item is a real, "
            "publicly-sourced example.) For a real-data report on actual current "
            "officials, run the program LIVE &#8212; the default; omit --offline "
            "&#8212; against data.ct.gov, SEEC, OSE, and cga.ct.gov.", synth))
        S.append(Spacer(1, 8))
    S.append(Paragraph(_esc(CAVEAT), note))
    S.append(Spacer(1, 8))

    # ========================================================================
    # COMPLETENESS VERDICT (the user's hard requirement: zero matches is only
    # acceptable if EVERY source was queried; otherwise label INCOMPLETE).
    # ========================================================================
    verdict = _completeness_verdict(coverage or {},
                                    getattr(municipal, "coverage", None) if municipal else None)
    vcolor = "#1e7a3c" if verdict["complete"] else "#7a1e1e"
    vlabel = ("COMPLETE — every required source was queried" if verdict["complete"]
              else "INCOMPLETE — not all required sources were queried")
    leads = leads or []
    TIER_COLOR = {"CONFIRMED": "#1e7a3c", "PROBABLE": "#1a5276",
                  "POSSIBLE": "#b8860b"}
    _dtier = display_tier
    findings_leads = [d for d in leads if d.get("confidence") in
                      ("CONFIRMED", "PROBABLE", "POSSIBLE")]
    if not findings_leads and leads:
        findings_leads = [d for d in leads if d.get("confidence") in
                          ("CONFIRMED", "PROBABLE", "POSSIBLE/REVIEW")]
    n_conf = sum(1 for d in findings_leads if d["confidence"] == "CONFIRMED")
    n_prob = sum(1 for d in findings_leads if d["confidence"] == "PROBABLE")
    n_poss = sum(1 for d in findings_leads if d["confidence"] == "POSSIBLE")
    n_muni = len(getattr(municipal, "known_findings", []) if municipal else [])
    conf_names = [d["person"] for d in findings_leads if d["confidence"] == "CONFIRMED"]
    # de-dup keep order
    conf_names = list(dict.fromkeys(conf_names))

    # ---- EXECUTIVE SUMMARY (lead with findings, not caveats) ---------------
    S.append(_H2("Executive Summary"))
    S.append(Paragraph(
        f"This {('LIVE' if mode == 'LIVE' else mode)} run resolved "
        f"<b>{n_conf} VERIFIED</b>, <b>{n_prob} HIGH PROBABILITY</b>, and "
        f"<b>{n_poss} POSSIBLE</b> legislator cannabis connection(s), plus "
        f"<b>{n_muni}</b> documented municipal connection(s). "
        + (f"Verified: <b>{_esc(', '.join(conf_names))}</b>. " if conf_names else "")
        + "Each finding below was actively relationship-resolved against public "
        "sources before a tier was assigned; surname coincidences with no "
        "relationship found are excluded. Investigation status: "
        f"<b><font color='{vcolor}'>{'COMPLETE' if verdict['complete'] else 'INCOMPLETE'}"
        f"</font></b> &#8212; a 'no match' is NOT a clearance; the sources not yet "
        f"queried are listed at the back (Coverage Gaps).", body))
    S.append(Spacer(1, 6))
    S.append(Paragraph(
        "<b>Confidence tiers</b> (assigned only after active resolution): "
        "<b><font color='#1e7a3c'>VERIFIED</font></b> &#8212; a primary source directly "
        "establishes the connection; "
        "<b><font color='#1a5276'>HIGH PROBABILITY</font></b> &#8212; multiple independent "
        "sources strongly indicate it; "
        "<b><font color='#b8860b'>POSSIBLE</font></b> &#8212; some evidence, not yet "
        "verified; "
        "<b>UNVERIFIED NAME MATCH</b> &#8212; only a name similarity, no relationship "
        "evidence found (excluded from the findings below).", note))
    S.append(Spacer(1, 10))

    _T_ORD = {"CONFIRMED": 0, "PROBABLE": 1, "POSSIBLE": 2, "POSSIBLE/REVIEW": 2}

    def _trim(s, n):
        s = (s or "").strip()
        return s if len(s) <= n else s[:n - 1].rstrip() + "…"

    _T_ORD = {"CONFIRMED": 0, "PROBABLE": 1, "POSSIBLE": 2, "POSSIBLE/REVIEW": 2}

    def _trim(s, n):
        s = (s or "").strip()
        return s if len(s) <= n else s[:n - 1].rstrip() + "…"

    def _findings_table(items):
        # One row PER OFFICIAL — merge that official's cannabis ties and de-dup the
        # evidence so the same person never repeats across near-identical rows.
        from collections import OrderedDict
        groups = OrderedDict()
        for d in items:
            groups.setdefault(d["person"], []).append(d)
        rows = []
        for person, ds in groups.items():
            ds = sorted(ds, key=lambda d: _T_ORD.get(d.get("confidence"), 9))
            top = ds[0]
            tier = top.get("confidence", "POSSIBLE")
            col = TIER_COLOR.get(tier, "#333333")
            kin = ("self" if top.get("same_first") and
                   top.get("name_similarity", 0) >= 92 else "self / relative")
            # all identity/relationship sources for this official, used for the inline
            # source link by the name AND the verification line. Web-evidence sources
            # (a page that actually NAMES the person) come first; the raw registry/
            # credential URL last. Only real http(s) links are kept (no dead anchors).
            srcs = []
            for d in ds:
                for u in (d.get("resolution", {}).get("sources", [])
                          + d.get("source_urls", [])):
                    if u and u.startswith("http") and u not in srcs:
                        srcs.append(u)
            # YEARS OF ACTIVE SERVICE (context the user asked for, every section).
            yrs = (top.get("years_served") or "").strip()
            served = (f"<br/>Served: <b>{_esc(yrs)}</b>" if yrs
                      else "<br/>Served: <i>(years not in dataset)</i>")
            # Source link RIGHT BY THE NAME. Only the VERIFIED tier asserts the identity
            # is certain; lower tiers link the source but say "verify".
            _lbl = ("verified identity source &#8599;" if tier == "CONFIRMED"
                    else "source &#8212; verify same person &#8599;")
            id_link = (f"<br/>{refs.link(srcs[0], _lbl)}"
                       if srcs else "<br/><font color='#7a1e1e'>no identity source "
                       "&#8212; UNVERIFIED</font>")
            official = (f"<b>{_esc(top['person'])}</b><br/>{_esc(top['role'])}{served}<br/>"
                        f"{_esc(top['party'])}, {_esc(top['district_or_town'])}<br/>"
                        f"<i>({kin})</i>{id_link}")
            ties = []
            for d in ds[:6]:
                ties.append(
                    f"<b>{_esc(_trim(d['cannabis_person'], 34))}</b> "
                    f"({_esc(d['cannabis_role'])}) &#8212; "
                    f"<i>{_esc(_trim(d['cannabis_entity'], 30))}</i>"
                    f"{(' [' + _esc(d['license_number']) + ']') if d.get('license_number') else ''}"
                    f"{(' reg ' + _esc(d['record_date'])) if d.get('record_date') else ''}")
            if len(ds) > 6:
                ties.append(f"(+{len(ds) - 6} more)")
            canna = "<br/>".join(ties)
            tcell = f"<b><font color='{col}'>{_esc(_dtier(tier))}</font></b>"
            # de-duplicated evidence across the official's ties
            seen, ev = set(), []
            for d in ds:
                for kind, etext, eurl in d.get("resolution", {}).get("evidence", []):
                    k = etext[:48]
                    if k in seen:
                        continue
                    seen.add(k)
                    ev.append((kind, etext, eurl))
            bits = [_esc(_trim(top.get("resolution", {}).get("explanation", ""), 240))]
            for kind, etext, eurl in ev[:3]:
                bits.append(f"&#8226; <b>{_esc(kind)}:</b> {_esc(_trim(etext, 150))}"
                            f"{refs.cite(eurl) if eurl else ''}")
            # VERIFICATION line on EVERY row: list the clickable sources, or state
            # plainly that none was located (no line item is left unverified-looking).
            if srcs:
                bits.append("<b>Verification (same-person sources):</b> "
                            + "".join(refs.cite(u) for u in srcs[:4]))
            else:
                bits.append("<b><font color='#7a1e1e'>Verification: NO PRIMARY SOURCE "
                            "LOCATED &#8212; treat as UNVERIFIED; confirm before "
                            "relying.</font></b>")
            assess = "<br/>".join(b for b in bits if b) or "&#8212;"
            rows.append([official, canna, tcell, assess])
        return _wrap_table(
            ["Official and office", "Cannabis ties (person / business / license)",
             "Tier", "Assessment, evidence and verification"],
            rows, [92, 116, 80, PAGE_W - 288], cell, cellh)

    def _by_tier(items):
        order = {"CONFIRMED": 0, "PROBABLE": 1, "POSSIBLE": 2, "POSSIBLE/REVIEW": 2}
        return sorted(items, key=lambda d: order.get(d.get("confidence"), 9))

    senators = _by_tier([d for d in findings_leads if "Senator" in d.get("role", "")])
    reps = _by_tier([d for d in findings_leads if "Representative" in d.get("role", "")])

    # ---- SECTION 1 — STATE SENATORS (past & present) + relatives -----------
    intro1 = Paragraph("Resolved connections only. Each was actively verified against "
                       "public sources (news, official/company bios, the business "
                       "registry, DCP credentials). A surname match with no relationship "
                       "found is NOT listed.", note)
    body1 = (_findings_table(senators) if senators else
             Paragraph("No resolved senator cannabis connection this run."
                       + ("" if verdict["complete"] else " (Run is INCOMPLETE — not a "
                          "clearance; see status at back.)"), small))
    _section(_H2("Section 1 &#8212; State Senators (Past &amp; Present) &amp; Relatives"),
             intro1, body1)
    S.append(Spacer(1, 8))

    # ---- SECTION 2 — STATE REPRESENTATIVES (past & present) + relatives -----
    body2 = (_findings_table(reps) if reps else
             Paragraph("No resolved representative cannabis connection this run."
                       + ("" if verdict["complete"] else " (INCOMPLETE — not a "
                          "clearance.)"), small))
    _section(_H2("Section 2 &#8212; State Representatives (Past &amp; Present) &amp; "
                 "Relatives"), body2)
    S.append(Spacer(1, 6))

    # ---- LEGISLATIVE VOTING & TIMELINE (per connected legislator) -----------
    voting_rows = []
    for d in findings_leads:
        vt = d.get("voting")
        if not vt:
            continue
        tl = "<br/>".join(f"{y} &#8212; {_esc(ev)}" for y, ev in vt.get("timeline", []))
        # ACTUAL cga.ct.gov roll-call votes (hard facts), then the web-sourced stance.
        rc_bits = []
        for rc in vt.get("rollcall", []) or []:
            vcol = ("#1e7a3c" if rc.get("vote") == "YEA" else
                    "#7a1e1e" if rc.get("vote") == "NAY" else "#666666")
            rc_bits.append(
                f"<b><font color='{vcol}'>{_esc(rc.get('vote'))}</font></b> on "
                f"{_esc(rc.get('bill'))} ({rc.get('year')}, {_esc(rc.get('era'))}; "
                f"floor {_esc(rc.get('tally'))})" + refs.cite(rc.get("url", "")))
        rec = vt.get("recusal") or {}
        rec_line = ""
        if rec.get("status"):
            rcol = ("#7a1e1e" if rec["status"] == "FOUND RECUSAL" else "#666666")
            rec_line = (f"<br/>Recusal: <font color='{rcol}'>{_esc(rec['status'])}"
                        f"</font>" + (refs.cite(*rec.get("sources", [])[:1])
                                      if rec.get("sources") else ""))
        stance_cell = (
            ("<br/>".join(rc_bits) + "<br/>" if rc_bits else "")
            + "<i>" + _esc(vt.get("stance", "undetermined")) + "</i>"
            + (refs.cite(*vt.get("sources", [])[:2]) if vt.get("sources") else "")
            + rec_line)
        voting_rows.append([
            f"<b>{_esc(d['person'])}</b><br/>{_esc(d['role'])}",
            _esc(", ".join(vt.get("eras", [])) or "verify"),
            stance_cell, tl or "&#8212;"])
    # de-dup by official
    seen_v, vrows = set(), []
    for r in voting_rows:
        if r[0] in seen_v:
            continue
        seen_v.add(r[0])
        vrows.append(r)
    if vrows:
        vt_tbl = _wrap_table(
            ["Legislator", "Cannabis era(s) in office",
             "Cannabis roll-call votes and stance",
             "Political vs cannabis timeline"],
            vrows, [104, 70, 156, PAGE_W - 330], cell, cellh)
        _section(_H2("Legislative Voting &amp; Cannabis Timeline"),
                 Paragraph("YEA/NAY are ACTUAL cga.ct.gov floor roll-call votes on the "
                           "landmark cannabis bills (2021 RERACA adult-use SB/HB 1201; "
                           "the 2012 medical act where available); the italic line is the "
                           "broader web-sourced stance. Recusal = a per-member cannabis "
                           "recusal/ethics check. The timeline shows whether the cannabis "
                           "interest arose before, during, or after service.", note),
                 vt_tbl)
        S.append(Spacer(1, 8))

    # ---- CAMPAIGN FINANCE (SEEC eCRIS) — cannabis money into legislators ----
    cf = campaign_finance or {}
    cf_recipients = cf.get("by_recipient", [])
    if cf_recipients:
        cf_total = cf.get("legislative_total", 0.0)
        cf_n = cf.get("legislative_count", 0)
        cf_intro = Paragraph(
            f"Contributions from CT cannabis operators, their executives/principals, and "
            f"employees to STATE LEGISLATIVE candidate committees, pulled live from the "
            f"SEEC eCRIS contribution search (by cannabis-employer and by contributor "
            f"name). <b>${cf_total:,.0f} across {cf_n} contribution(s)</b> to "
            f"{len(cf_recipients)} committee(s). A contribution is a lawful, disclosed "
            f"donation &#8212; it is context for the appearance of access, NOT an "
            f"allegation. Recipients are matched to a cannabis-era legislator by committee "
            f"name + district where possible; unmatched committees are shown as "
            f"[committee].", note)
        cf_rows = []
        for g in cf_recipients[:40]:
            who = g.get("legislator") or g.get("recipient", "")
            head = (f"<b>{_esc(who)}</b>"
                    + (f"<br/>{_esc(g.get('office'))}" if g.get("office") else "")
                    + (f", Dist. {_esc(g.get('district'))}" if g.get("district") else "")
                    + (f"<br/>{_esc(g.get('party'))}" if g.get("party") else ""))
            donors = "; ".join(_trim(d, 28) for d in g.get("donors", [])[:6])
            emps = "; ".join(_trim(e, 26) for e in g.get("employers", [])[:5])
            ties = (f"{_esc(donors)}"
                    + (f"<br/><i>{_esc(emps)}</i>" if emps else "")
                    + (f"<br/>{_esc(', '.join(g.get('years', [])[:6]))}"
                       if g.get("years") else ""))
            amt = (f"<b>${g.get('total', 0):,.0f}</b><br/>{g.get('n', 0)} gift(s)")
            src = "".join(refs.cite(u) for u in g.get("sources", [])[:2])
            note_cell = ("Lawful disclosed contribution; verify identity and recency "
                         "in eCRIS." + src)
            cf_rows.append([head, ties, amt, note_cell])
        cf_tbl = _wrap_table(
            ["Recipient committee / legislator", "Cannabis donor(s) (employer and years)",
             "Amount", "Note and source"],
            cf_rows, [120, 150, 60, PAGE_W - 330], cell, cellh)
        cap = cf.get("capped", {}) or {}
        cf_caveat = ""
        if cap.get("employers") or cap.get("contributors"):
            cf_caveat = (f" Coverage note: {cap.get('employers', 0)} employer + "
                         f"{cap.get('contributors', 0)} contributor searches were over "
                         f"budget and NOT run this cycle (re-run to extend).")
        _section(_H2("Campaign Finance &#8212; Cannabis-Linked Contributions (SEEC eCRIS)"),
                 cf_intro, cf_tbl,
                 Paragraph("Source: SEEC eCRIS public contribution search "
                           "(seec.ct.gov/eCrisReporting). Only state-legislative "
                           "recipients are shown; gubernatorial/other recipients of "
                           "cannabis money are excluded from this section." + cf_caveat,
                           small))
        S.append(Spacer(1, 8))
    elif cf:
        _section(_H2("Campaign Finance &#8212; Cannabis-Linked Contributions (SEEC eCRIS)"),
                 Paragraph("No cannabis-linked contribution to a state legislative "
                           "committee surfaced this run"
                           + (" (eCRIS searched live; none matched)." if mode == "LIVE"
                              else " (offline fixture)."), note))
        S.append(Spacer(1, 8))

    # ---- CANNABIS LOBBYING (CT Office of State Ethics) ---------------------
    lob = lobbying or {}
    lob_roster = lob.get("roster", [])
    lob_matches = lob.get("legislator_matches", [])
    if lob_roster or lob_matches:
        lob_intro = Paragraph(
            f"Registered cannabis-industry lobbyist communicators from the CT Office of "
            f"State Ethics roster (data.ct.gov dataset 4ixq-tnwe): "
            f"<b>{lob.get('cannabis_lobbyist_count', 0)} communicator(s)</b> across "
            f"<b>{lob.get('org_count', 0)} organization(s)</b>. A registered lobbyist is "
            f"a lawful, disclosed role &#8212; this maps the cannabis influence roster "
            f"and flags any overlap with a legislator's name. NOTE: contract lobbying "
            f"firms hired BY a cannabis client (client&#8594;firm registrations) live on "
            f"the OSE eLobbyist portal, not this bulk dataset, so they are not yet "
            f"captured unless the organization itself is cannabis-named.", note)
        flowables = [lob_intro]
        # (a) the highest-signal output: legislator name overlaps
        if lob_matches:
            mrows = []
            for m in lob_matches[:30]:
                col = ("#7a1e1e" if m.get("same_person") else
                       "#1a5276" if m.get("same_first") else "#b8860b")
                tier = ("LEGISLATOR IS A CANNABIS LOBBYIST" if m.get("same_person")
                        else "SAME NAME — VERIFY" if m.get("same_first")
                        else "SURNAME ONLY — VERIFY")
                mrows.append([
                    f"<b>{_esc(m.get('legislator'))}</b>"
                    + (f"<br/>{_esc(m.get('party'))}, Dist. {_esc(str(m.get('district')))}"
                       if m.get("district") else ""),
                    f"<b>{_esc(m.get('communicator'))}</b><br/>lobbies for "
                    f"<i>{_esc(m.get('organization'))}</i>",
                    f"<b><font color='{col}'>{_esc(tier)}</font></b>",
                    _esc(m.get("note", "")) + refs.cite(m.get("source", ""))])
            flowables.append(_wrap_table(
                ["Legislator", "Cannabis lobbyist (organization)", "Flag",
                 "Assessment and source"],
                mrows, [110, 150, 90, PAGE_W - 350], cell, cellh, head_bg="#7a1e1e"))
        # (b) the cannabis lobbyist roster (context)
        if lob_roster:
            rrows = []
            for g in lob_roster[:40]:
                comms = "; ".join(_trim(c, 26) for c in g.get("communicators", [])[:8])
                rrows.append([
                    f"<b>{_esc(g.get('organization'))}</b>",
                    _esc(comms),
                    _esc(", ".join(g.get("cities", [])[:4]))
                    + (f"<br/>{_esc(', '.join(g.get('years', [])[:4]))}"
                       if g.get("years") else "")
                    + refs.cite(g.get("source", ""))])
            flowables.append(Paragraph("<b>Cannabis lobbying organizations and their "
                                       "registered communicators:</b>", small))
            flowables.append(_wrap_table(
                ["Organization", "Registered communicators", "City / year (source)"],
                rrows, [150, PAGE_W - 320, 170], cell, cellh))
        # DONATIONS tied to these cannabis-lobby orgs — giver, recipient, amount, date
        # (joined from the SEEC eCRIS contributions by employer/organization name).
        cf_rows = (campaign_finance or {}).get("rows", [])
        org_names = [g.get("organization", "") for g in lob_roster]

        def _org_for(employer):
            e = (employer or "").lower()
            for o in org_names:
                for tok in (o or "").lower().split():
                    if len(tok) >= 5 and tok not in ("cannabis", "chamber", "commerce",
                                                     "connecticut") and tok in e:
                        return o
            return ""
        dono = [(_org_for(r.get("employer")), r) for r in cf_rows]
        dono = [(o, r) for o, r in dono if o]
        if dono:
            drows = []
            for o, r in dono[:40]:
                drows.append([
                    f"<b>{_esc(r.get('donor'))}</b><br/><i>{_esc(o)}</i>",
                    f"{_esc(r.get('legislator') or r.get('committee'))}"
                    + (f"<br/>{_esc(r.get('office'))}"
                       + (f", Dist. {_esc(str(r.get('district')))}"
                          if r.get('district') else "") if r.get('office') else ""),
                    f"<b>${r.get('amount', 0):,.0f}</b><br/>{_esc(r.get('date'))}"
                    + refs.cite(r.get('source_url', ''))])
            flowables.append(Paragraph(
                "<b>Disclosed contributions linked to these cannabis-lobby "
                "organizations</b> (giver &#8594; recipient, amount, date &#8212; from "
                "SEEC eCRIS; lawful, publicly disclosed):", small))
            flowables.append(_wrap_table(
                ["Giver (and organization)", "Recipient",
                 "Amount and date (source)"],
                drows, [160, PAGE_W - 330, 170], cell, cellh))
        _section(_H2("Cannabis Lobbying &amp; Money (CT OSE + SEEC)"), *flowables)
        S.append(Spacer(1, 8))

    # ---- SECTION 3 — MUNICIPAL LEADERS + CONNECTIONS -----------------------
    intro3 = Paragraph("The Simsbury First Selectman &#8594; Pullman &amp; Comley "
                       "cannabis attorney &#8594; Curaleaf siting case is the WORKED "
                       "EXAMPLE of a category &#8212; a town official, their family, or "
                       "the town's counsel tied to cannabis &#8212; now GENERALIZED to "
                       "every host town: a host-town roster (every town with an operating "
                       "cannabis business), per-town official cannabis-tie leads (&#167;3a), "
                       "and town-attorney cannabis chains (&#167;3b). Town-official and "
                       "town-counsel rosters are not bulk-published, so coverage is by "
                       "live search and remains partial; absence is not a clearance.", note)
    known = getattr(municipal, "known_findings", []) if municipal else []
    muni_dossier = []
    if municipal is not None:
        for dossier in municipal.dossiers:
            for c in dossier.connections:
                if c.classification in ("CONFIRMED", "UNCONFIRMED") and \
                        c.subject_kind in ("official", "spouse/family", "firm"):
                    muni_dossier.append((dossier, c))
    if known or muni_dossier:
        mrows = []
        for k in known:
            col = TIER_COLOR.get(k.get("tier", "POSSIBLE"), "#333333")
            official = (f"<b>{_esc(k.get('official'))}</b><br/>{_esc(k.get('role'))}"
                        f"<br/>{_esc(k.get('town'))}")
            conn = (f"{_esc(k.get('relation') or 'tie')} &#8594;<br/>"
                    f"<b>{_esc(k.get('connected_person'))}</b><br/>cannabis: "
                    f"<i>{_esc(k.get('cannabis_business'))}</i>")
            tcell = (f"<b><font color='{col}'>{_esc(_dtier(k.get('tier')))}</font></b>"
                     + ("<br/>appearance concern" if k.get("appearance_concern") else ""))
            srcs = "".join(refs.cite(u) for u in k.get("sources", []))
            assess = _esc(k.get("explanation")) + (f"<br/>Sources:{srcs}" if srcs else "")
            mrows.append([official, conn, tcell, assess])
        for dossier, c in muni_dossier:
            col = TIER_COLOR.get(c.classification, "#b8860b")
            mrows.append([
                f"<b>{_esc(c.subject_name)}</b><br/>{_esc(dossier.town)}",
                f"cannabis: <i>{_esc(dossier.operator)}</i>",
                f"<b><font color='{col}'>{_esc(c.classification)}</font></b>",
                _esc(c.explanation) + "".join(refs.cite(u) for u in c.citations)])
        mbody = _wrap_table(
            ["Official and town", "Connection", "Tier", "Assessment and sources"],
            mrows, [96, 132, 56, PAGE_W - 284], cell, cellh)
    else:
        mbody = Paragraph("No documented municipal connection in this run's data yet. "
                          "Town-level official / family / counsel records have no "
                          "statewide bulk API (INCOMPLETE); systematic per-town "
                          "resolution is the top open item.", small)
    _section(_H2("Section 3 &#8212; Municipal Leaders &amp; Their Connections"),
             intro3, mbody)
    S.append(Spacer(1, 8))

    # ---- §3a MUNICIPAL OFFICIAL CANNABIS TIES (Glassman category, generalized) ----
    ot_find = getattr(municipal, "official_tie_findings", []) if municipal else []
    if ot_find:
        otrows = []
        for f in ot_find[:40]:
            otrows.append([
                f"<b>{_esc(f.get('town'))}</b>",
                f"{_esc(_trim(f.get('headline', ''), 120))}<br/>"
                f"<i>{_esc(_trim(f.get('snippet', ''), 200))}</i>",
                f"<font color='#b8860b'><b>POSSIBLE &#8212; web lead, VERIFY</b></font>"
                + refs.cite(f.get("source", ""))])
        _section(_H3("Municipal Official Cannabis Ties (Glassman Category, Generalized)"),
                 Paragraph("The Simsbury / Mary Glassman pattern &#8212; a town official "
                           "(or their family) tied to cannabis &#8212; applied across "
                           "every host town by live search. These are POSSIBLE LEADS that "
                           "name a town-leadership role alongside a cannabis tie in a "
                           "credible result; each MUST be verified against the primary "
                           "source before relying. Absence here is not a clearance "
                           "(town-official rosters are not bulk-published).", note),
                 _wrap_table(["Town", "Surfaced result", "Status and source"],
                             otrows, [80, PAGE_W - 290, 130], cell, cellh,
                             head_bg="#7a1e1e"))
        S.append(Spacer(1, 8))

    # ---- §3b TOWN-ATTORNEY CANNABIS CHAINS --------------------------------
    ta_find = getattr(municipal, "town_attorney_findings", []) if municipal else []
    if ta_find:
        tarows = []
        for f in ta_find:
            disc = f.get("discovered")
            col = "#b8860b" if disc else "#7a1e1e"
            flag = ("POSSIBLE cannabis-practice town counsel (web-discovered LEAD "
                    "&#8212; VERIFY the firm is this town's counsel)" if disc else
                    "cannabis-practice town counsel (sourced)")
            tarows.append([
                f"<b>{_esc(f.get('town'))}</b>",
                f"<b>{_esc(f.get('firm'))}</b>"
                + (f"<br/><i>{_esc(f.get('cannabis_lead'))}</i>"
                   if f.get("cannabis_lead") else ""),
                _esc(f.get("cannabis_practice", "")),
                f"<font color='{col}'><b>{flag}</b></font> &#8212; appearance concern "
                f"for cannabis matters before this town."
                + "".join(refs.cite(u) for u in f.get("sources", [])[:3])])
        _section(_H3("Town-Attorney Cannabis Chains"),
                 Paragraph("Host towns whose town counsel / corporation counsel is a "
                           "law firm with a cannabis practice (the firm advises the town "
                           "AND represents cannabis interests). This is an appearance "
                           "concern, not a substantial conflict by itself.", note),
                 _wrap_table(["Town", "Town-counsel firm", "Cannabis practice",
                              "Why it matters and source"],
                             tarows, [80, 130, 110, PAGE_W - 320], cell, cellh,
                             head_bg="#7a1e1e"))
        S.append(Spacer(1, 8))

    # ---- §3c HOST-TOWN ROSTER (every town with an operating business) ------
    roster = getattr(municipal, "host_town_roster", []) if municipal else []
    if roster:
        rrows = []
        for g in roster[:80]:
            zon = g.get("zoning") or "&#8212;"
            zcol = ("#7a1e1e" if any(k in (g.get("zoning") or "").lower()
                                     for k in ("morator", "prohibit")) else "#16412b")
            counsel = (f"<b>{_esc(g.get('counsel'))}</b> "
                       f"<font color='#7a1e1e'>(cannabis-practice)</font>"
                       if g.get("cannabis_counsel") else "<i>not identified</i>")
            rrows.append([
                f"<b>{_esc(g.get('town'))}</b>",
                _esc("; ".join(_trim(o, 28) for o in g.get("operators", [])[:6])),
                f"<font color='{zcol}'>{_esc(g.get('zoning') or '—')}</font>",
                counsel])
        n_cc = sum(1 for g in roster if g.get("cannabis_counsel"))
        _section(_H3("Host-Town Roster"),
                 Paragraph(f"All {len(roster)} CT town(s) hosting an operating cannabis "
                           f"business this run, with zoning status and (where identified) "
                           f"town counsel. {n_cc} town(s) have a cannabis-practice firm as "
                           f"counsel; the rest were not identified (town-counsel rosters "
                           f"are not bulk-published).", note),
                 _wrap_table(["Town", "Operating cannabis business(es)",
                              "Zoning status", "Town counsel"],
                             rrows, [90, PAGE_W - 360, 80, 190], cell, cellh))
        S.append(Spacer(1, 8))

    # ---- SIDE NOTE — vote recusals (deliberately minor) --------------------
    officials = {}
    for d in findings_leads:
        officials.setdefault(d['person'], d)
    rec_names = ", ".join(sorted(officials)) if officials else "none surfaced"
    S.append(_H3("Side Note &#8212; Cannabis Votes &amp; Recusals"))
    S.append(Paragraph(
        "Recusals are a WEAK signal: officials frequently do NOT recuse from cannabis "
        "votes they arguably should, so a missing recusal proves little and is treated "
        "here only as a side note. CT cannabis eras: medical 2012 (PA 12-55), adult-use "
        "2021 (RERACA). Officials with a connection above whose cannabis votes/recusals "
        "warrant a manual look: " + _esc(rec_names) + ". (CGA roll-call/recusal records "
        "are not yet integrated &#8212; INCOMPLETE.)", small))
    S.append(Spacer(1, 8))

    # ---- Municipal cannabis policy (zoning) — moratorium reconciliation -----
    zoning = getattr(municipal, "zoning", None) if municipal else None
    if zoning:
        host = {}
        for dd in (municipal.dossiers if municipal else []):
            host.setdefault((dd.town or "").lower(), []).append(dd.operator)
        def _restrictive(s):
            s = (s or "").lower()
            return "morator" in s or "prohibit" in s
        conflict_towns = [z for z in zoning if _restrictive(z.get("status"))
                          and (z.get("town") or "").lower() in host]
        appr = sum(1 for z in zoning if "approv" in (z.get("status") or "").lower())
        moro = sum(1 for z in zoning if "morator" in (z.get("status") or "").lower())
        proh = sum(1 for z in zoning if "prohibit" in (z.get("status") or "").lower())
        S.append(_H3("Municipal Cannabis Policy (Zoning) &#8212; Context"))
        S.append(Paragraph(
            f"{len(zoning)} towns recorded (Approved {appr} · Moratorium {moro} · "
            f"Prohibited {proh}). A zoning status governs NEW licenses only — it does "
            f"not remove a pre-existing operation, so a restrictive town can still host "
            f"a grandfathered business.", note))
        if conflict_towns:
            crows = [[f"<b>{_esc(z.get('town'))}</b>",
                      f"<font color='#7a1e1e'><b>{_esc(z.get('status'))}</b></font>",
                      _esc("; ".join(host.get((z.get('town') or '').lower(), [])))]
                     for z in sorted(conflict_towns, key=lambda z: z.get("town") or "")]
            S.append(Paragraph(f"<b>{len(conflict_towns)} restrictive town(s) that still "
                               f"host an operating cannabis business:</b>", small))
            S.append(_wrap_table(
                ["Town", "Zoning status", "Operating business(es)"],
                crows, [120, 90, PAGE_W - 210], cell, cellh, head_bg="#7a1e1e"))
        S.append(Spacer(1, 8))


    # -- run summary table (counts MUST match the findings above) ----------
    S.append(_H2("Run Summary &amp; Self-Validation"))
    n_known = len(getattr(municipal, "known_findings", []) if municipal else [])
    actual_conf = sum(1 for d in findings_leads if d.get("confidence") == "CONFIRMED")
    actual_prob = sum(1 for d in findings_leads if d.get("confidence") == "PROBABLE")
    actual_poss = sum(1 for d in findings_leads if d.get("confidence") == "POSSIBLE")
    keys = [("legislators", "Legislators in roster"),
            ("cross_referenced", "Cross-referenced (cannabis era 2012+)"),
            ("cannabis_persons", "Cannabis credential holders / principals resolved"),
            ("confirmed_findings", "VERIFIED legislator findings"),
            ("probable_findings", "HIGH PROBABILITY legislator findings"),
            ("possible_findings", "POSSIBLE legislator findings"),
            ("senator_findings", "State Senator findings"),
            ("representative_findings", "State Representative findings"),
            ("vote_review_candidates", "Vote-review candidates"),
            ("cannabis_contributions",
             "Cannabis-linked legislative contributions (SEEC)"),
            ("cannabis_lobbyists", "Cannabis-industry lobbyists (OSE)"),
            ("cannabis_lobbyist_leg_matches",
             "Cannabis lobbyist <-> legislator name matches")]
    data = [["Metric", "Count"]] + [[lbl, str(counts.get(k, 0))] for k, lbl in keys]
    if counts.get("cannabis_contribution_total"):
        data.append(["Cannabis-linked contribution total (SEEC)",
                     f"${counts.get('cannabis_contribution_total', 0):,.0f}"])
    data.append(["Municipal findings", str(n_known + len(muni_dossier))])
    t = Table(data, colWidths=[340, 70], hAlign="LEFT")
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16412b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f6f4")]),
    ]))
    S.append(t)
    # Self-validation: summary counts vs what is actually rendered above.
    warns = []
    if counts.get("confirmed_findings", 0) != actual_conf:
        warns.append(f"VERIFIED count {counts.get('confirmed_findings')} != "
                     f"{actual_conf} rendered")
    if counts.get("probable_findings", 0) != actual_prob:
        warns.append("HIGH PROBABILITY count mismatch")
    if counts.get("possible_findings", 0) != actual_poss:
        warns.append("POSSIBLE count mismatch")
    if (counts.get("senator_findings", 0) + counts.get("representative_findings", 0)) \
            != (len(senators) + len(reps)):
        warns.append("chamber split mismatch")
    S.append(Paragraph(
        ("<b><font color='#1e7a3c'>Self-validation PASSED</font></b> — summary counts "
         "match the findings rendered above." if not warns else
         "<b><font color='#7a1e1e'>Self-validation WARNING:</font></b> " +
         _esc("; ".join(warns))), small))

    # -- COVERAGE GAPS: required sources not fully queried (moved to the back) --
    if not verdict["complete"]:
        S.append(_H2("Coverage Gaps &#8212; Required Sources Not Fully Queried"))
        S.append(Paragraph("A 'no match' on any person is NOT a clearance. The sources "
                           "below would confirm or refute connections and are not yet "
                           "fully integrated:", small))
        miss_rows = [[_esc(k), _esc(d), _esc(n)] for k, d, n in verdict["missing"]]
        S.append(_wrap_table(["Required source NOT fully queried", "What it would add",
                              "Why / status"], miss_rows,
                             [150, 150, PAGE_W - 300], cell, cellh, head_bg="#7a1e1e"))
        S.append(Spacer(1, 8))

    # -- coverage (which sources were actually queried) -------------------
    if coverage:
        S.append(_H2("Sources Queried (Coverage)"))
        S.append(Paragraph("Absence of a match is &#8216;no match found,&#8217; "
                           "<b>not</b> proof of no involvement. Sources marked NOT "
                           "QUERIED were not exhaustively checked this run.", note))
        merged = dict(coverage)
        if municipal is not None and getattr(municipal, "coverage", None):
            for k, v in municipal.coverage.items():
                merged[f"(town) {k}"] = v
        cov = [["Source", "Status", "Records"]]
        for label, c in merged.items():
            cov.append([label, _STATUS_PLAIN.get(c["status"], c["status"]),
                        str(c["count"])])
        ct = Table(cov, colWidths=[250, 200, 50], hAlign="LEFT")
        ct.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a5276")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#f3f6f4")]),
        ]))
        S.append(ct)
        gaps = [(lbl, c) for lbl, c in merged.items()
                if c["status"] in ("unavailable", "disabled") and c.get("note")]
        if gaps:
            S.append(Paragraph("Coverage gaps — why a source could not be "
                               "exhaustively queried:", small))
            for lbl, c in gaps:
                S.append(Paragraph(f"<font size=7><b>{_esc(lbl)}</b> — "
                                   f"{_esc(c['note'])}</font>", small))

        # Live datasets queried — clickable source links (so the report always
        # carries live, clickable URLs even when no individual finding was made).
        live_urls: list[tuple[str, str]] = []
        if legislators:
            u = legislators[0].provenance.source_url
            if u:
                live_urls.append(("Legislators (current + historical)", u))
        if municipal is not None and municipal.dossiers:
            for u in sorted({d.facility.provenance.source_url
                             for d in municipal.dossiers if d.facility.provenance.source_url}):
                live_urls.append(("Cannabis facilities / licenses", u))
        if live_urls:
            S.append(Paragraph("Live datasets queried (clickable):", small))
            for lbl, u in live_urls:
                S.append(Paragraph(f"<font size=8>{_esc(lbl)}: "
                                   f"{refs.link(u)}{refs.cite(u)}</font>", small))

    # -- legal standard ---------------------------------------------------
    S.append(_H2("Legal Standard &amp; Methodology"))
    S.append(Paragraph(_esc(LEGAL_PREAMBLE), small))

    # (Findings now live in Sections 1–3 above; recusals are a side note there.
    # The legacy per-Finding / per-town-dossier blocks are intentionally removed —
    # no surname coincidences, no passive facility lists, no duplicate sections.)

    # -- references appendix (numbered, clickable) ------------------------
    S.append(PageBreak())
    S.append(_H2("References"))
    S.append(Paragraph("Every citation marker above links to its source; the same "
                       "sources are listed here with live, clickable URLs.", small))
    S.append(Spacer(1, 4))
    if refs.order:
        for url in refs.order:
            n = refs.index[url]
            S.append(Paragraph(f"[{n}] {refs.link(url)}", small))
            S.append(Spacer(1, 1))
    else:
        S.append(Paragraph("No sources cited this run.", small))

    doc.build(S, onFirstPage=_page_footer, onLaterPages=_page_footer)


def write_municipal_review_queue(path: Path, rows: list[dict]) -> None:
    cols = ["town", "operator", "subject", "subject_kind", "connection_type",
            "classification", "confidence", "is_private_individual",
            "substantial_conflict", "explanation", "source_url"]
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in cols})


def write_all(result, cfg: dict, municipal=None, pdf_path: str | None = None,
              report_number: int | None = None) -> dict:
    out = cfg["output"]
    run_date = date.today().isoformat()
    paths = {
        "tracker": Path(out["tracker_xlsx"]),
        "findings_md": Path(out["findings_md"]),
        "findings_pdf": Path(pdf_path) if pdf_path else Path(out["findings_pdf"]),
        "review_queue": Path(out["review_queue_csv"]),
    }
    if municipal is not None:
        paths["municipal_review_queue"] = Path(
            out.get("municipal_review_queue_csv", "out/municipal_review_queue.csv"))
    for p in paths.values():
        p.parent.mkdir(parents=True, exist_ok=True)
    write_tracker(paths["tracker"], result.legislators, result.findings, run_date,
                  municipal=municipal)
    write_review_queue(paths["review_queue"], result.review_rows)
    cov = getattr(result, "coverage", {})
    mode = getattr(result, "mode", "OFFLINE")
    write_findings_md(paths["findings_md"], result.legislators, result.findings,
                      result.recusals, run_date, result.counts, municipal=municipal,
                      coverage=cov, mode=mode,
                      leads=getattr(result, "legislator_cannabis_leads", []),
                      network=getattr(result, "network", None))
    if municipal is not None:
        write_municipal_review_queue(paths["municipal_review_queue"],
                                     municipal.review_rows)
    try:
        write_findings_pdf(paths["findings_pdf"], result.findings, result.recusals,
                           run_date, result.counts, legislators=result.legislators,
                           municipal=municipal, coverage=cov, mode=mode,
                           report_number=report_number,
                           leads=getattr(result, "legislator_cannabis_leads", []),
                           network=getattr(result, "network", None),
                           campaign_finance=getattr(result, "campaign_finance", {}),
                           lobbying=getattr(result, "lobbying", {}))
    except Exception as e:  # noqa: BLE001 — PDF is best-effort; never block the run
        paths["findings_pdf"] = None  # type: ignore
        print(f"[warn] PDF generation skipped: {e}")
    return {k: (str(v) if v else None) for k, v in paths.items()}


def finalize_report(result, cfg: dict, municipal=None,
                    push_to_downloads: bool = True) -> dict:
    """Assign the next persistent report number, write the PDF as
    reports/CTCannabisPoliticalCheck_<N>.pdf (never overwriting a prior report),
    copy it to the TOP of ~/Downloads, and write the other outputs to out/.
    Returns {number, report_pdf, downloads_pdf, paths}."""
    n = next_report_number()
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    report_pdf = REPORTS_DIR / f"{PROGRAM_NAME}_{n}.pdf"
    paths = write_all(result, cfg, municipal=municipal, pdf_path=str(report_pdf),
                      report_number=n)
    downloads_pdf = None
    if push_to_downloads and report_pdf.exists():
        dl = Path.home() / "Downloads" / f"{PROGRAM_NAME}_{n}.pdf"
        try:
            shutil.copy2(report_pdf, dl)
            downloads_pdf = str(dl)
        except Exception as e:  # noqa: BLE001
            print(f"[warn] could not copy report to Downloads: {e}")
    _record_report_number(n, {
        "date": date.today().isoformat(),
        "mode": getattr(result, "mode", "OFFLINE"),
        "report_pdf": str(report_pdf),
        "downloads_pdf": downloads_pdf,
        "legislators": result.counts.get("legislators"),
        "cross_referenced": result.counts.get("cross_referenced"),
        "host_towns": (municipal.counts.get("host_towns") if municipal else 0),
    })
    return {"number": n, "report_pdf": str(report_pdf),
            "downloads_pdf": downloads_pdf, "paths": paths}
